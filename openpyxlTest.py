#!/usr/bin/env python
# -*- coding: utf-8 -*-
import openpyxl

from openpyxl import Workbook
from openpyxl.cell import Cell
from openpyxl.styles import Font, Side, Border, Alignment
import base64

import sys
import os
import cx_Oracle



reload(sys)
sys.setdefaultencoding('utf-8')
os.environ['LANG'] = 'en_US.UTF-8'

# Workbook 생성
wb = Workbook()
# Workbook을 생성하면 적어도 하나의 워크시트를 생성한다.
ws1 = wb.active
ws1.title = 'Example1'

con = cx_Oracle.Connection('DIGIUSER/DIGIUSER@REALDIGI')
cursor = con.cursor()

qry = '''
SELECT
        ROWNUM RNUM
        , A.*
FROM (
        SELECT A.CP_BRCD
              , B.EPSD_NO
              , B.TTLE
              , B.TTLE_ENCODE
              , A.WRKNM
              , A.BRCD 
              , B.UPDT_DTIME  
        FROM LCPLATFORM.LCP_PRDT A, LCPLATFORM.LCP_EPSD B 
        WHERE A.BRCD= B.BRCD 
        AND A.CP_CD = 'E0100' 
        AND A.PBLSERL_STT ='00027' 
        AND B.EPSD_STT ='00047'  
        AND (TO_CHAR(B.UPDT_DTIME,'YYYYMMDD') = '20180524' OR TO_CHAR(B.RGST_DTIME,'YYYYMMDD')  = '20180524')  
        ORDER BY B.UPDT_DTIME, A.CP_BRCD, B.EPSD_NO
    ) A 
'''
cursor.execute(qry)

s = '한글'
#print str(unicode(s))


column_num = 2
row = cursor.fetchone()
while row:
    column_char = 'a'
    for x in range(1, 8):
        #print(row[x - 1].decode('cp949'))
        #print(row[2].decode('cp949'))
        row_index = row[0] +1

        ws1.cell(row=row_index, column=1).value = row[1]
        ws1.cell(row=row_index, column=2).value = row[2]
        ws1.cell(row=row_index, column=3).value = row[3].decode('cp949')
        ws1.cell(row=row_index, column=4).value = base64.decodestring(row[4])
        ws1.cell(row=row_index, column=5).value = row[5].decode('cp949')
        ws1.cell(row=row_index, column=6).value = row[6]
        ws1.cell(row=row_index, column=7).value = row[7]


        #ws1[column_char + str(column_num)] = row[2].decode('cp949')
        column_char = chr(ord(column_char) + 1)

    column_num = column_num + 1
    row = cursor.fetchone()

#print con.version
con.close()



# 1.시트의 행렬의 번호를 입력하여 Text 입력이 가능
#ws1['A1'] = 'Hello'

#ws1.cell(row=3, column=3).value = 777

# 2-1 시트에서 셀을 변수로 만들고
#c = ws1['A2']

# 2-2 아래에서처럼 각 셀에 대해 value 입력, font 설정 등을 할 수 있다.
#c.value = "A2 zone"
#c.font =Font(name='Arial', size=14)
#c.border = Border(left=Side(border_style="thin", color='FF000000'), right=Side(border_style="thin", color='FF000000'), top=Side(border_style="thin", color='FF000000'), bottom=Side(border_style="thin", color='FF000000'))
#c.alignment = Alignment(horizontal='center', vertical='center')

# 저장
wb.save('test.xlsx')
wb.close()
